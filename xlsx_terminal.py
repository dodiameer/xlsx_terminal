//Importing Required Modules
import sys
from openpyxl import Workbook, load_workbook

if len(sys.argv) > 1: //Check if input contained options
    if sys.argv[1] == "new": //Create new file with input file name
        wb = Workbook()
        ws = wb.active
        wb.save(sys.argv[2])
    if sys.argv[1] == "add": //Add input data to specified file
        wb = load_workbook(sys.argv[2])
        ws = wb.active
        active_row = int(sys.argv[3])
        for i in range(4, len(sys.argv)):
            ws.cell(active_row, i-3, str(sys.argv[i]))
        wb.save(sys.argv[2])
    if sys.argv[1] == "remove": //Remove specified row's data
        wb = load_workbook(sys.argv[2])
        ws = wb.active
        active_row = int(sys.argv[3])
        ws.delete_rows(active_row)
        wb.save(sys.argv[2])
    else: //Print error if option is invalid
        print('Error: Invalid Option!')
else: //Print error if no options were put in by user
    print('Please input an option and try again!')
